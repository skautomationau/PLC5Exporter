import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import win32com.client as win32com
import os
import re
import threading
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class PLC5ExcelExporter:
    def __init__(self, root):
        self.root = root
        self.root.title("PLC-5 RSP to Excel Exporter")
        self.root.geometry("700x500")
        
        # Folder containing one or more RSP files
        self.rsp_folder = None
        # Optional separate output folder
        self.output_folder = None
        self.is_processing = False
        
        self.setup_ui()
    
    def setup_ui(self):
        # File selection frame
        file_frame = ttk.LabelFrame(self.root, text="File Selection", padding=10)
        file_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(file_frame, text="RSP Folder:").grid(row=0, column=0, sticky="w", pady=5)
        self.file_label = ttk.Label(file_frame, text="No folder selected", foreground="gray")
        self.file_label.grid(row=0, column=1, sticky="w", padx=10)
        ttk.Button(file_frame, text="Browse...", command=self.browse_rsp).grid(row=0, column=2, padx=5)
        
        ttk.Label(file_frame, text="Output Folder:").grid(row=1, column=0, sticky="w", pady=5)
        self.output_label = ttk.Label(file_frame, text="Same as RSP folder", foreground="gray")
        self.output_label.grid(row=1, column=1, sticky="w", padx=10)
        ttk.Button(file_frame, text="Browse...", command=self.browse_output).grid(row=1, column=2, padx=5)
        
        # Recursive search option
        self.recursive = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            file_frame,
            text="Recursive subfolders",
            variable=self.recursive
        ).grid(row=2, column=1, sticky="w", pady=5)
        
        # Export options
        options_frame = ttk.LabelFrame(self.root, text="Export Options", padding=10)
        options_frame.pack(fill="x", padx=10, pady=5)
        
        self.export_tags = tk.BooleanVar(value=True)
        self.export_timers = tk.BooleanVar(value=True)
        self.export_counters = tk.BooleanVar(value=True)
        self.export_controls = tk.BooleanVar(value=True)
        self.export_arrays = tk.BooleanVar(value=True)
        self.export_messages = tk.BooleanVar(value=True)
        self.export_io = tk.BooleanVar(value=True)
        self.export_rungs = tk.BooleanVar(value=True)
        self.export_datatable = tk.BooleanVar(value=False)
        
        ttk.Checkbutton(options_frame, text="Tags/Addresses", variable=self.export_tags).grid(row=0, column=0, sticky="w")
        ttk.Checkbutton(options_frame, text="Timers", variable=self.export_timers).grid(row=0, column=1, sticky="w")
        ttk.Checkbutton(options_frame, text="Counters", variable=self.export_counters).grid(row=0, column=2, sticky="w")
        ttk.Checkbutton(options_frame, text="Controls", variable=self.export_controls).grid(row=1, column=0, sticky="w")
        ttk.Checkbutton(options_frame, text="Arrays", variable=self.export_arrays).grid(row=1, column=1, sticky="w")
        ttk.Checkbutton(options_frame, text="Messages", variable=self.export_messages).grid(row=1, column=2, sticky="w")
        ttk.Checkbutton(options_frame, text="I/O Points", variable=self.export_io).grid(row=2, column=0, sticky="w")
        ttk.Checkbutton(options_frame, text="Ladder Rungs", variable=self.export_rungs).grid(row=2, column=1, sticky="w")
        ttk.Checkbutton(options_frame, text="Data Table (slow)", variable=self.export_datatable).grid(row=2, column=2, sticky="w")
        
        # Progress frame
        progress_frame = ttk.LabelFrame(self.root, text="Progress", padding=10)
        progress_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress_bar.pack(fill="x", pady=5)
        
        self.log_text = tk.Text(progress_frame, height=10, wrap="word", state="disabled")
        self.log_text.pack(fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(self.log_text, command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # Action buttons
        button_frame = ttk.Frame(self.root)
        button_frame.pack(fill="x", padx=10, pady=10)
        
        self.export_btn = ttk.Button(button_frame, text="Export to Excel", command=self.start_export, state="disabled")
        self.export_btn.pack(side="left", padx=5)
        
        ttk.Button(button_frame, text="Exit", command=self.root.quit).pack(side="right", padx=5)
    
    def browse_rsp(self):
        folder = filedialog.askdirectory(
            title="Select Folder Containing PLC-5 RSP Files"
        )
        if folder:
            self.rsp_folder = folder
            self.file_label.config(text=self.rsp_folder, foreground="black")
            if not self.output_folder:
                self.output_folder = self.rsp_folder
                self.output_label.config(text=self.output_folder, foreground="black")
            self.export_btn.config(state="normal")
            self.log(f"Selected folder: {self.rsp_folder}")
    
    def browse_output(self):
        folder = filedialog.askdirectory(title="Select Output Folder")
        if folder:
            self.output_folder = folder
            self.output_label.config(text=folder, foreground="black")
            self.log(f"Output folder: {folder}")
    
    def log(self, message):
        self.log_text.config(state="normal")
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")
        self.root.update()
    
    def start_export(self):
        if self.is_processing:
            return
        
        if not self.rsp_folder:
            messagebox.showerror("Error", "Please select an RSP folder")
            return
        
        self.is_processing = True
        self.export_btn.config(state="disabled")
        self.progress_bar.start()
        
        thread = threading.Thread(target=self.export_data, daemon=True)
        thread.start()
    
    def export_data(self):
        rslogix5 = None
        try:
            self.log("Opening RSLogix5 Application...")
            rslogix5 = win32com.Dispatch("RSLogix5.Application.5")
            rslogix5.visible = True
            
            base_folder = os.path.abspath(self.rsp_folder)
            self.log(f"Scanning folder: {base_folder}")
            
            # Build list of .rsp files
            rsp_files = []
            if self.recursive.get():
                for root_dir, _, files in os.walk(base_folder):
                    for f in files:
                        if f.lower().endswith(".rsp"):
                            rsp_files.append(os.path.join(root_dir, f))
            else:
                for f in os.listdir(base_folder):
                    full_path = os.path.join(base_folder, f)
                    if os.path.isfile(full_path) and f.lower().endswith(".rsp"):
                        rsp_files.append(full_path)
            
            if not rsp_files:
                messagebox.showerror("Error", "No .rsp files found in the selected folder")
                self.log("No .rsp files found.")
                return
            
            self.log(f"Found {len(rsp_files)} RSP file(s).")
            if not self.output_folder:
                self.output_folder = base_folder
                self.output_label.config(text=self.output_folder, foreground="black")
            
            total = len(rsp_files)
            
            for idx, rsp_path in enumerate(rsp_files, start=1):
                self.log("=" * 60)
                self.log(f"Processing {idx}/{total}: {rsp_path}")
                
                try:
                    abs_path = os.path.abspath(rsp_path)
                    self.log(f"Opening project: {abs_path}")
                    project = rslogix5.FileOpen(abs_path, False, False, True)
                    
                    program_files = project.ProgramFiles
                    addr_sym_records = project.AddrSymRecords
                    datafiles = project.DataFiles
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    base_name = os.path.splitext(os.path.basename(rsp_path))[0]
                    
                    # Create workbook
                    wb = Workbook()
                    wb.remove(wb.active)  # Remove default sheet
                    
                    # Collect all data
                    self.log("  Analyzing ladder logic...")
                    data_collection = self.analyze_ladder_logic(program_files, addr_sym_records, datafiles)
                    
                    # Add rungs if requested
                    if self.export_rungs.get():
                        self.log("  Extracting ladder rungs...")
                        data_collection['rungs'] = self.collect_rungs(program_files)
                    
                    # Add datatable if requested
                    if self.export_datatable.get():
                        self.log("  Extracting data table...")
                        data_collection['datatable'] = self.collect_datatable(datafiles)
                    
                    # Write all sheets to Excel
                    self.log("  Writing Excel file...")
                    self.write_excel_workbook(wb, data_collection)
                    
                    # Save workbook
                    excel_file = os.path.join(
                        self.output_folder,
                        f"{base_name}_Export_{timestamp}.xlsx"
                    )
                    wb.save(excel_file)
                    self.log(f"  File saved: {excel_file}")
                    
                    # Close this project before moving to the next
                    project.Close(False)
                
                except Exception as per_file_exc:
                    import traceback
                    self.log(f"ERROR processing {rsp_path}: {per_file_exc}")
                    self.log(traceback.format_exc())
                    # Continue with the next file
                    continue
            
            self.log("=" * 60)
            self.log("All exports completed.")
            messagebox.showinfo("Success", f"Export completed for {len(rsp_files)} RSP file(s).")
        
        except Exception as e:
            import traceback
            self.log(f"ERROR (global): {str(e)}")
            self.log(traceback.format_exc())
            messagebox.showerror("Error", f"Export failed:\n{str(e)}")
        finally:
            if rslogix5 is not None:
                try:
                    self.log("Closing RSLogix5...")
                    rslogix5.Quit(True, False)
                except Exception:
                    pass
            
            self.is_processing = False
            self.progress_bar.stop()
            self.export_btn.config(state="normal")
    
    def analyze_ladder_logic(self, program_files, addr_sym_records, datafiles):
        """Collect all data from ladder analysis"""
        data = {
            'timers': {},
            'counters': {},
            'controls': {},
            'arrays': [],
            'messages': {},
            'tags': {},
            'io': {}
        }
        
        for file_idx in range(2, program_files.Count()):
            try:
                ladder_file = program_files(file_idx)
                if ladder_file and ladder_file.NumberOfRungs() > 0:
                    self.log(f"  Analyzing: {ladder_file.Name}")
                    for rung_idx in range(ladder_file.NumberOfRungs()):
                        try:
                            rung_ascii = ladder_file.GetRungAsAscii(rung_idx)
                            
                            # Extract addresses
                            self.extract_addresses(
                                rung_ascii, data['tags'], data['io'], 
                                addr_sym_records, datafiles
                            )
                            
                            # Extract components
                            if self.export_timers.get():
                                self.extract_timers(rung_ascii, data['timers'], addr_sym_records)
                            if self.export_counters.get():
                                self.extract_counters(rung_ascii, data['counters'], addr_sym_records)
                            if self.export_controls.get():
                                self.extract_controls(rung_ascii, data['controls'], addr_sym_records)
                            if self.export_messages.get():
                                self.extract_messages(rung_ascii, data['messages'], addr_sym_records)
                        except Exception:
                            continue
            except Exception:
                continue
        
        return data
    
    def extract_addresses(self, rung, tags, io, addr_sym_records, datafiles):
        """Extract addresses from rung"""
        pattern = re.compile(r'\b([IONBFLTCRS]:\d+(?:/\d+)?|[BNTCR]\d+:\d+(?:/\d+)?)\b')
        
        for addr in pattern.findall(rung):
            if addr in tags or addr in io:
                continue
            
            symbol, desc = self.get_symbol_desc(addr, addr_sym_records)
            value = self.get_value(addr, datafiles)
            
            if addr.startswith('I:') or addr.startswith('O:'):
                io[addr] = {
                    'Type': 'Input' if addr.startswith('I:') else 'Output',
                    'Address': addr,
                    'Symbol': symbol,
                    'Description': desc,
                    'Value': value
                }
            else:
                tags[addr] = {
                    'Address': addr,
                    'Symbol': symbol,
                    'Description': desc,
                    'DataType': self.get_data_type(addr),
                    'Value': value
                }
    
    def extract_timers(self, rung, timers, addr_sym_records):
        pattern = re.compile(r'(TON|TOF|RTO)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)')
        for match in pattern.findall(rung):
            addr = match[1]
            if addr not in timers:
                symbol, desc = self.get_symbol_desc(addr, addr_sym_records)
                timers[addr] = {
                    'Type': match[0],
                    'Address': addr,
                    'Symbol': symbol,
                    'Description': desc,
                    'Base': match[2],
                    'PRE': match[3],
                    'ACC': match[4]
                }
    
    def extract_counters(self, rung, counters, addr_sym_records):
        pattern = re.compile(r'(CTU|CTD)\s+(\S+)\s+(\S+)\s+(\S+)')
        for match in pattern.findall(rung):
            addr = match[1]
            if addr not in counters:
                symbol, desc = self.get_symbol_desc(addr, addr_sym_records)
                counters[addr] = {
                    'Type': match[0],
                    'Address': addr,
                    'Symbol': symbol,
                    'Description': desc,
                    'PRE': match[2],
                    'ACC': match[3]
                }
    
    def extract_controls(self, rung, controls, addr_sym_records):
        pattern = re.compile(r'(FAL|FSC|FFL|FFU|COP|DDT|FBC)\s+(\S+)\s+(\S+)\s+(\S+)')
        for match in pattern.findall(rung):
            addr = match[1]
            if addr not in controls:
                symbol, desc = self.get_symbol_desc(addr, addr_sym_records)
                controls[addr] = {
                    'Instruction': match[0],
                    'Address': addr,
                    'Symbol': symbol,
                    'Description': desc,
                    'Length': match[2],
                    'Position': match[3]
                }
    
    def extract_messages(self, rung, messages, addr_sym_records):
        """
        Extract MSG instructions from a rung.

        We look for each occurrence of 'MSG' in the ASCII rung, grab everything
        up to 'EOR' (if present) or the end of the string, then split into
        whitespace-separated parameters.

        Expected PLC-5 MSG parameter convention:

            MSG <ControlBlock> <PLCFamily> <DataType> <Direction>
                <LocalAddr> <LocalLength> <RemoteNode> <RemoteAddr> <RemoteLength> <PortType> <Channel>
        """
        pos = 0
        while True:
            idx = rung.find('MSG', pos)
            if idx == -1:
                break

            # Ensure 'MSG' is a standalone token
            before_ok = (idx == 0) or not rung[idx - 1].isalnum()
            after_ok = (idx + 3 >= len(rung)) or rung[idx + 3].isspace()
            if not (before_ok and after_ok):
                pos = idx + 3
                continue

            # First space after MSG = start of parameters
            param_start = rung.find(' ', idx + 3)
            if param_start == -1:
                break

            # End at ' EOR' if present, else end of string
            end_idx = rung.find(' EOR', idx)
            if end_idx == -1:
                end_idx = len(rung)

            block = rung[param_start + 1:end_idx].strip()
            pos = end_idx

            if not block:
                continue

            parts = block.split()
            if len(parts) < 4:
                # Need at least: ControlBlock, PLC_Family, DataType, Direction
                continue

            def get(i: int) -> str:
                return parts[i] if i < len(parts) else ""

            # Map parameters
            ctrl_block  = get(0)
            plc_family  = get(1)
            data_type   = get(2)
            direction   = get(3)

            local_addr  = get(4)
            local_len   = get(5)
            remote_node = get(6)
            remote_addr = get(7)
            remote_len  = get(8)
            port_type   = get(9)
            channel     = get(10)

            if not ctrl_block or ctrl_block in messages:
                continue

            symbol, desc = self.get_symbol_desc(ctrl_block, addr_sym_records)

            msg_entry = {
                # Core identification
                'Address': ctrl_block,
                'Symbol': symbol,
                'Description': desc,

                # Semantic PLC-5 style fields
                'PLC_Family': plc_family,
                'DataType': data_type,
                'Direction': direction,
                'LocalAddr': local_addr,
                'LocalLength': local_len,
                'RemoteNode': remote_node,
                'RemoteAddr': remote_addr,
                'RemoteLength': remote_len,
                'PortType': port_type,
                'Channel': channel,

                # Excel heading aliases requested:
                # LocalLength  -> Size
                # RemoteNode   -> PortNumber
                # RemoteLength -> DHPlusNode
                'Size': local_len,
                'PortNumber': remote_node,
                'DHPlusNode': remote_len,

                # Raw parameters (handy for debugging)
                'RawParameters': ' '.join(parts),
            }

            # Backward compatible generic names (old version used these)
            if 'Type' not in msg_entry:
                msg_entry['Type'] = plc_family
            if 'ThisPLC' not in msg_entry:
                msg_entry['ThisPLC'] = data_type
            if 'Length' not in msg_entry:
                msg_entry['Length'] = direction
            if 'Port' not in msg_entry:
                msg_entry['Port'] = local_addr
            if 'Target' not in msg_entry:
                msg_entry['Target'] = local_len
            if 'Node' not in msg_entry:
                msg_entry['Node'] = remote_node

            messages[ctrl_block] = msg_entry
    
    def collect_rungs(self, program_files):
        """Collect all rungs"""
        rungs = []
        for file_idx in range(2, program_files.Count()):
            try:
                ladder_file = program_files(file_idx)
                if ladder_file and ladder_file.NumberOfRungs() > 0:
                    for rung_idx in range(ladder_file.NumberOfRungs()):
                        try:
                            rungs.append({
                                'File_Name': ladder_file.Name,
                                'File_Number': ladder_file.FileNumber,
                                'Rung_Number': rung_idx,
                                'Rung_ASCII': ladder_file.GetRungAsAscii(rung_idx)
                            })
                        except Exception:
                            continue
            except Exception:
                continue
        return rungs
    
    def collect_datatable(self, datafiles):
        """Collect datatable values"""
        data = []
        for file_idx in range(datafiles.Count()):
            try:
                datafile = datafiles(file_idx)
                if datafile:
                    file_type = datafile.TypeAsString
                    if file_type in ['B', 'N', 'F', 'L', 'T', 'C', 'R']:
                        file_num = datafile.FileNumber
                        for elem in range(min(datafile.NumberOfElements, 1000)):
                            addr = f"{file_type}{file_num}:{elem}"
                            try:
                                data.append({
                                    'FileType': file_type,
                                    'FileNumber': file_num,
                                    'Element': elem,
                                    'Address': addr,
                                    'Value': datafiles.GetDataValue(addr)
                                })
                            except Exception:
                                pass
            except Exception:
                continue
        return data
    
    def write_excel_workbook(self, wb, data):
        """Write all data to Excel sheets"""
        if self.export_tags.get() and data['tags']:
            self.write_sheet(wb, 'Tags', 
                           ['Address', 'Symbol', 'Description', 'DataType', 'Value'],
                           data['tags'].values())
            self.log(f"  Wrote {len(data['tags'])} tags")
        
        if self.export_io.get() and data['io']:
            self.write_sheet(wb, 'IO',
                           ['Type', 'Address', 'Symbol', 'Description', 'Value'],
                           data['io'].values())
            self.log(f"  Wrote {len(data['io'])} I/O points")
        
        if self.export_timers.get() and data['timers']:
            self.write_sheet(wb, 'Timers',
                           ['Type', 'Address', 'Symbol', 'Description', 'Base', 'PRE', 'ACC'],
                           data['timers'].values())
            self.log(f"  Wrote {len(data['timers'])} timers")
        
        if self.export_counters.get() and data['counters']:
            self.write_sheet(wb, 'Counters',
                           ['Type', 'Address', 'Symbol', 'Description', 'PRE', 'ACC'],
                           data['counters'].values())
            self.log(f"  Wrote {len(data['counters'])} counters")
        
        if self.export_controls.get() and data['controls']:
            self.write_sheet(wb, 'Controls',
                           ['Instruction', 'Address', 'Symbol', 'Description', 'Length', 'Position'],
                           data['controls'].values())
            self.log(f"  Wrote {len(data['controls'])} controls")
        
        if self.export_messages.get() and data['messages']:
            self.write_sheet(
                wb,
                'Messages',
                [
                    'Address',
                    'Symbol',
                    'Description',
                    'PLC_Family',
                    'DataType',
                    'Direction',
                    'LocalAddr',
                    'Size',         # was LocalLength
                    'PortNumber',   # was RemoteNode
                    'RemoteAddr',
                    'DHPlusNode',   # was RemoteLength
                    'PortType',
                    'Channel',
                    'RawParameters',
                ],
                data['messages'].values()
            )
            self.log(f"  Wrote {len(data['messages'])} messages")
     
        if 'rungs' in data and data['rungs']:
            self.write_sheet(wb, 'Rungs',
                           ['File_Name', 'File_Number', 'Rung_Number', 'Rung_ASCII'],
                           data['rungs'])
            self.log(f"  Wrote {len(data['rungs'])} rungs")
        
        if 'datatable' in data and data['datatable']:
            self.write_sheet(wb, 'DataTable',
                           ['FileType', 'FileNumber', 'Element', 'Address', 'Value'],
                           data['datatable'])
            self.log(f"  Wrote {len(data['datatable'])} datatable values")
    
    def write_sheet(self, wb, sheet_name, headers, rows):
        """Write data to a sheet efficiently, expanding Description into Desc1..Desc5 when present"""
        ws = wb.create_sheet(title=sheet_name)

        # If a Description column exists, replace it with Desc1..Desc5
        if 'Description' in headers:
            desc_index = headers.index('Description')
            new_headers = (
                headers[:desc_index]
                + ['Desc1', 'Desc2', 'Desc3', 'Desc4', 'Desc5']
                + headers[desc_index + 1:]
            )
        else:
            new_headers = headers

        ws.append(new_headers)

        for row in rows:
            if isinstance(row, dict):
                values = [row.get(h, '') for h in headers]
            else:
                values = list(row)

            # Split description into 5 columns if applicable
            if 'Description' in headers:
                desc_index = headers.index('Description')
                desc_text = str(values[desc_index]).strip()
                # Split on pipe, comma, or semicolon — keep non-empty, trimmed parts
                split_desc = [part.strip() for part in re.split(r'\||,|;', desc_text) if part.strip()]
                desc_parts = (split_desc + [''] * 5)[:5]  # Always 5 columns
                values = values[:desc_index] + desc_parts + values[desc_index + 1:]

            ws.append(values)

        # Auto-size columns
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    def get_symbol_desc(self, addr, addr_sym_records):
        try:
            record = addr_sym_records.GetRecordViaAddrOrSym(addr, 0)
            if record:
                return (
                    record.Symbol or "", 
                    (record.Description or "").replace('\r\n', ' | ')
                )
        except Exception:
            pass
        return "", ""
    
    def get_value(self, addr, datafiles):
        try:
            return datafiles.GetDataValue(addr)
        except Exception:
            return ""
    
    def get_data_type(self, addr):
        if ':' not in addr:
            return "Unknown"
        prefix = addr.split(':')[0].replace('#', '')
        types = {
            'I': 'Input',
            'O': 'Output',
            'B': 'Bit',
            'N': 'Integer',
            'F': 'Float',
            'L': 'Long',
            'T': 'Timer',
            'C': 'Counter',
            'R': 'Control',
            'S': 'Status'
        }
        return types.get(prefix, prefix)


if __name__ == "__main__":
    root = tk.Tk()
    app = PLC5ExcelExporter(root)
    root.mainloop()
